# export.py - Versión con gráficos centrados y colores azul institucional

import io
import os
import logging
from datetime import datetime
from flask import Blueprint, request, send_file
import sqlite3

from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                Paragraph, Spacer, PageBreak)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.barcharts import VerticalBarChart

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, Reference, Series

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

export_bp = Blueprint('export', __name__)

if os.environ.get('RENDER'):
    DATABASE = '/tmp/elecciones.db'
else:
    DATABASE = 'elecciones.db'

# Colores institucionales
AZUL_PRINCIPAL = '#0B3B6E'
AZUL_SECUNDARIO = '#1B5E9E'
AMARILLO = '#F5A623'

RL_AZUL_PRINCIPAL = colors.HexColor(AZUL_PRINCIPAL)
RL_AZUL_SECUNDARIO = colors.HexColor(AZUL_SECUNDARIO)
RL_AMARILLO = colors.HexColor(AMARILLO)

def _fetch_data(modules=None):
    db = sqlite3.connect(DATABASE)
    db.row_factory = sqlite3.Row
    data = {}
    try:
        if modules is None or 'secretarias' in modules:
            rows = db.execute("SELECT name, empleados, votos_reportados FROM secretarias ORDER BY name").fetchall()
            data['secretarias'] = [dict(r) for r in rows]
        if modules is None or 'institutos' in modules:
            rows = db.execute("SELECT name, empleados, votos_reportados FROM institutos ORDER BY name").fetchall()
            data['institutos'] = [dict(r) for r in rows]
        if modules is None or 'jubilados' in modules:
            rows = db.execute("SELECT name, total, votos_reportados FROM jubilados ORDER BY name").fetchall()
            data['jubilados'] = [dict(r) for r in rows]
    except Exception as e:
        logger.error(f"Error fetching data: {e}")
    finally:
        db.close()
    return data

def fill_excel_sheet(ws, rows, titulo, campo_total):
    if not rows:
        logger.warning(f"No hay datos para {titulo}, no se crea hoja")
        return

    ws.title = titulo[:31]
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor=AZUL_PRINCIPAL[1:])
    center = Alignment(horizontal='center')

    ws['A1'] = f"CONSULTA POPULAR NACIONAL 2026 - {titulo.upper()}"
    ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws['A1'].fill = PatternFill("solid", fgColor=AZUL_PRINCIPAL[1:])
    ws.merge_cells('A1:F1')

    ws['A2'] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws.merge_cells('A2:F2')

    headers = ['#', 'NOMBRE', 'TOTAL', 'VOTARON', 'FALTAN', '% PART.']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center

    total_emp = total_vot = 0
    start_row = 5
    for idx, row in enumerate(rows, 1):
        emp = row[campo_total]
        vot = row['votos_reportados']
        total_emp += emp
        total_vot += vot
        faltan = emp - vot
        pct = f"{(vot/emp*100):.1f}%" if emp > 0 else "0%"
        ws.cell(row=idx+4, column=1, value=idx).alignment = center
        ws.cell(row=idx+4, column=2, value=row['name']).alignment = Alignment(wrap_text=True)
        ws.cell(row=idx+4, column=3, value=emp).alignment = center
        ws.cell(row=idx+4, column=4, value=vot).alignment = center
        ws.cell(row=idx+4, column=5, value=faltan).alignment = center
        ws.cell(row=idx+4, column=6, value=pct).alignment = center
    end_row = idx + 4

    last = end_row + 1
    ws.cell(row=last, column=2, value="TOTAL").font = Font(bold=True)
    ws.cell(row=last, column=3, value=total_emp).font = Font(bold=True)
    ws.cell(row=last, column=4, value=total_vot).font = Font(bold=True)
    ws.cell(row=last, column=5, value=total_emp - total_vot).font = Font(bold=True)
    pct_g = f"{(total_vot/total_emp*100):.1f}%" if total_emp else "0%"
    ws.cell(row=last, column=6, value=pct_g).font = Font(bold=True)

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 50
    for c in ['C','D','E','F']:
        ws.column_dimensions[c].width = 13

    if len(rows) > 1:
        try:
            chart = BarChart()
            chart.title = f"Participación por {titulo}"
            chart.y_axis.title = "Cantidad"
            chart.x_axis.title = "Entidades"
            chart.legend.position = 'b'
            chart.width = 25
            chart.height = 10

            data_vot = Reference(ws, min_col=4, min_row=start_row, max_row=end_row)
            data_falt = Reference(ws, min_col=5, min_row=start_row, max_row=end_row)
            cats = Reference(ws, min_col=2, min_row=start_row, max_row=end_row)

            series_vot = Series(data_vot, title="Votaron")
            series_falt = Series(data_falt, title="Faltan")
            chart.append(series_vot)
            chart.append(series_falt)
            chart.set_categories(cats)

            ws.add_chart(chart, f"H{last + 3}")
            logger.info(f"Gráfico añadido en hoja {titulo}")
        except Exception as e:
            logger.error(f"Error creando gráfico Excel para {titulo}: {e}")

@export_bp.route('/export/excel')
def export_excel():
    modules = request.args.getlist('modules')
    if not modules:
        modules = ['secretarias', 'institutos', 'jubilados']

    data = _fetch_data(modules)
    if not data:
        return "No hay datos para exportar", 404

    wb = openpyxl.Workbook()
    default = wb.active
    wb.remove(default)

    if 'secretarias' in data and data['secretarias']:
        ws = wb.create_sheet("Secretarías")
        fill_excel_sheet(ws, data['secretarias'], "Secretarías", "empleados")
    if 'institutos' in data and data['institutos']:
        ws = wb.create_sheet("Institutos")
        fill_excel_sheet(ws, data['institutos'], "Institutos", "empleados")
    if 'jubilados' in data and data['jubilados']:
        ws = wb.create_sheet("Jubilados")
        fill_excel_sheet(ws, data['jubilados'], "Jubilados", "total")

    if not wb.worksheets:
        return "No hay datos para los módulos seleccionados", 404

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"participacion_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

def create_bar_chart(data_rows, width=450, height=150):
    if len(data_rows) == 0:
        return None

    names = [r[0][:20] for r in data_rows]
    votos = [r[2] for r in data_rows]
    faltan = [r[1] - r[2] for r in data_rows]

    margin_left = (width - 380) // 2
    if margin_left < 0:
        margin_left = 10

    drawing = Drawing(width, height)
    bc = VerticalBarChart()
    bc.x = margin_left
    bc.y = 30
    bc.width = 380
    bc.height = height - 60
    bc.data = [votos, faltan]
    bc.strokeColor = colors.black
    bc.valueAxis.valueMin = 0
    bc.valueAxis.valueMax = max(r[1] for r in data_rows) * 1.1
    bc.categoryAxis.categoryNames = names
    bc.categoryAxis.labels.boxAnchor = 'ne'
    bc.categoryAxis.labels.dx = 8
    bc.categoryAxis.labels.dy = -2
    bc.categoryAxis.labels.angle = 45
    bc.categoryAxis.labels.fontSize = 7
    bc.bars[0].fillColor = RL_AZUL_PRINCIPAL   # Azul para Votaron
    bc.bars[1].fillColor = RL_AMARILLO         # Amarillo para Faltan
    bc.bars[0].strokeColor = colors.black
    bc.bars[1].strokeColor = colors.black
    drawing.add(bc)
    return drawing

@export_bp.route('/export/pdf')
def export_pdf():
    modules = request.args.getlist('modules')
    if not modules:
        modules = ['secretarias', 'institutos', 'jubilados']

    data = _fetch_data(modules)
    if not data:
        return "No hay datos para exportar", 404

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)

    styles = getSampleStyleSheet()
    wrap_s  = ParagraphStyle('W', parent=styles['Normal'], fontSize=7, leading=9, alignment=TA_LEFT)
    ctr_s   = ParagraphStyle('C', parent=styles['Normal'], fontSize=7, leading=9, alignment=TA_CENTER)
    bold_s  = ParagraphStyle('B', parent=styles['Normal'], fontSize=7, leading=9, alignment=TA_CENTER, fontName='Helvetica-Bold')
    hdr_s   = ParagraphStyle('H', parent=styles['Normal'], fontSize=8, leading=10, alignment=TA_CENTER, fontName='Helvetica-Bold', textColor=colors.white)
    title_s = ParagraphStyle('T', parent=styles['Heading1'], fontSize=16, textColor=RL_AZUL_PRINCIPAL, alignment=TA_CENTER, fontName='Helvetica-Bold')
    sub_s   = ParagraphStyle('S', parent=styles['Normal'], fontSize=9, textColor=RL_AZUL_SECUNDARIO, alignment=TA_CENTER, spaceAfter=14)

    story = []
    story.append(Paragraph("CONSULTA POPULAR NACIONAL 2026", title_s))
    story.append(Paragraph(f"Gobernación del Estado Bolívar — {datetime.now().strftime('%d/%m/%Y %H:%M')}", sub_s))

    def make_table(rows, campo_label):
        headers = [Paragraph(h, hdr_s) for h in ['#', 'NOMBRE', campo_label.upper(), 'VOTARON', 'FALTAN', '%']]
        tdata = [headers]
        t_emp = t_vot = 0
        for idx, row in enumerate(rows, 1):
            emp = row[1]
            vot = row[2]
            t_emp += emp
            t_vot += vot
            pct = (vot/emp*100) if emp > 0 else 0
            tdata.append([
                Paragraph(str(idx), ctr_s),
                Paragraph(str(row[0]), wrap_s),
                Paragraph(str(emp), ctr_s),
                Paragraph(str(vot), ctr_s),
                Paragraph(str(emp-vot), ctr_s),
                Paragraph(f"{pct:.1f}%", ctr_s),
            ])
        pct_g = (t_vot/t_emp*100) if t_emp else 0
        tdata.append([Paragraph('', ctr_s), Paragraph('<b>TOTAL</b>', wrap_s),
                      Paragraph(f'<b>{t_emp}</b>', bold_s), Paragraph(f'<b>{t_vot}</b>', bold_s),
                      Paragraph(f'<b>{t_emp-t_vot}</b>', bold_s), Paragraph(f'<b>{pct_g:.1f}%</b>', bold_s)])
        t = Table(tdata, colWidths=[1.2*cm, 9.5*cm, 2.2*cm, 2.2*cm, 2.2*cm, 2.2*cm], repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), RL_AZUL_PRINCIPAL),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('LEFTPADDING', (0,0), (-1,-1), 4), ('RIGHTPADDING', (0,0), (-1,-1), 4),
            ('TOPPADDING', (0,0), (-1,-1), 5), ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ('GRID', (0,0), (-1,-2), 0.5, colors.grey),
            ('BOX', (0,0), (-1,-1), 1, RL_AZUL_PRINCIPAL),
            ('ROWBACKGROUNDS', (0,1), (-1,-2), [colors.white, colors.HexColor('#F4F6FB')]),
            ('BACKGROUND', (0,-1), (-1,-1), RL_AMARILLO),
            ('LINEABOVE', (0,-1), (-1,-1), 2, RL_AZUL_PRINCIPAL),
        ]))
        return t

    first = True
    for module in ['secretarias', 'institutos', 'jubilados']:
        if module in data and data[module]:
            if not first:
                story.append(PageBreak())
            titulo = module.capitalize()
            story.append(Paragraph(titulo, ParagraphStyle('sec', parent=styles['Heading2'], fontSize=12, textColor=RL_AZUL_PRINCIPAL, alignment=TA_CENTER)))
            if module == 'jubilados':
                rows = [(r['name'], r['total'], r['votos_reportados']) for r in data[module]]
                campo_label = "Total"
            else:
                rows = [(r['name'], r['empleados'], r['votos_reportados']) for r in data[module]]
                campo_label = "Empleados"
            story.append(make_table(rows, campo_label))
            if len(rows) > 1:
                try:
                    chart = create_bar_chart(rows)
                    if chart:
                        story.append(Spacer(1, 0.5*cm))
                        chart_table = Table([[chart]], colWidths=[doc.width])
                        chart_table.setStyle(TableStyle([
                            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                        ]))
                        story.append(chart_table)
                except Exception as e:
                    logger.error(f"Error creando gráfico PDF para {module}: {e}")
            first = False

    if not first:
        doc.build(story)
        buf.seek(0)
        fname = f"participacion_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
        return send_file(buf, as_attachment=True, download_name=fname, mimetype='application/pdf')
    else:
        return "No hay datos para los módulos seleccionados", 404